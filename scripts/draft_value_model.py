#!/usr/bin/env python3
"""
Draft-value curve model fitted to historical Win Shares.

Goal: build an empirical draft-pick value curve from ~22 years of outcomes and
compare it to the hand-chosen VALUE_CURVE used in the app
(app/stash-value/ValueScaleClient.tsx). Either it justifies our curve, or it
tells us where ours is wrong.

Source: public/Draft Valuation.xlsx
  - One sheet per draft year (2003-2024), each with career stats per pick.
  - Columns (0-indexed): Pk=1, WS=18, BPM=20, VORP=21.

Key modeling decisions (see README block in the printed output):
  * Metric = career Win Shares (WS). Standard basis for draft-value work and
    matches how this app treats picks: long-horizon tradeable assets.
  * Maturity bias: a 2003 class has 22 seasons of accumulation; 2024 has ~1.
    Raw career WS therefore can't be averaged across years directly. We convert
    each pick to its SHARE of its own draft class's total WS, then average that
    share across years. Shares are scale-free, so every class informs the SHAPE
    of the curve equally regardless of age.
  * Negative career WS floored at 0 (a bust is a worthless asset, not negative).
  * We fit smooth monotone forms (exponential, power-law, hybrid) to the
    averaged-share curve and pick the best by R^2, then normalize pick 1 = 1.0.
"""

import json
import math
from pathlib import Path

import numpy as np
import openpyxl
from scipy.optimize import curve_fit

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "public" / "Draft Valuation.xlsx"
OUT_JSON = ROOT / "public" / "sim-output" / "draft_value_model.json"

PK_COL, WS_COL, BPM_COL, VORP_COL = 1, 18, 20, 21
N_PICKS = 60

# The hand-chosen curve currently in the app (normalized pick 1 = 1.0).
APP_CURVE = {
    1: 1.0, 2: 0.84007, 3: 0.7516, 4: 0.6895, 5: 0.64183,
    6: 0.60223, 7: 0.56745, 8: 0.53702, 9: 0.50913, 10: 0.48013,
    11: 0.45108, 12: 0.42505, 13: 0.4007, 14: 0.37892, 15: 0.35868,
    16: 0.34123, 17: 0.32595, 18: 0.31275, 19: 0.29943, 20: 0.28758,
    21: 0.27428, 22: 0.26237, 23: 0.24967, 24: 0.2383, 25: 0.22705,
    26: 0.21682, 27: 0.20703, 28: 0.19608, 29: 0.18613, 30: 0.17563,
    31: 0.15018, 32: 0.14385, 33: 0.13697, 34: 0.1311, 35: 0.125,
    36: 0.11983, 37: 0.11443, 38: 0.1097, 39: 0.10373, 40: 0.0994,
    41: 0.0951, 42: 0.09053, 43: 0.08663, 44: 0.0824, 45: 0.07825,
    46: 0.0741, 47: 0.06998, 48: 0.0659, 49: 0.06215, 50: 0.05815,
    51: 0.05445, 52: 0.05108, 53: 0.04745, 54: 0.0435, 55: 0.03955,
    56: 0.0369, 57: 0.03333, 58: 0.02947, 59: 0.02625, 60: 0.02242,
}


def load_picks(metric_col=WS_COL):
    """Return dict: year -> {pick -> career metric (negatives floored at 0)}."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    data = {}
    for name in wb.sheetnames:
        if not name.isdigit():
            continue
        ws = wb[name]
        year = int(name)
        picks = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                pk = int(row[PK_COL])
            except (TypeError, ValueError, IndexError):
                continue
            if not (1 <= pk <= N_PICKS):
                continue
            mv = row[metric_col] if len(row) > metric_col else None
            if not isinstance(mv, (int, float)):
                continue
            picks[pk] = max(0.0, float(mv))  # floor busts at 0
        if picks:
            data[year] = picks
    return data


def per_slot_curves(data):
    """
    Build two per-slot aggregations across years:
      share_mean[p] : mean within-year value-share at slot p (maturity-neutral)
      raw_mean[p]   : naive mean career WS at slot p (shows the maturity bias)
    Also return per-slot sample counts.
    """
    share_by_slot = {p: [] for p in range(1, N_PICKS + 1)}
    raw_by_slot = {p: [] for p in range(1, N_PICKS + 1)}

    for year, picks in data.items():
        total = sum(picks.values())
        if total <= 0:
            continue
        for p, wsv in picks.items():
            share_by_slot[p].append(wsv / total)
            raw_by_slot[p].append(wsv)

    share_mean = np.array([
        np.mean(share_by_slot[p]) if share_by_slot[p] else 0.0
        for p in range(1, N_PICKS + 1)
    ])
    share_median = np.array([
        np.median(share_by_slot[p]) if share_by_slot[p] else 0.0
        for p in range(1, N_PICKS + 1)
    ])
    raw_mean = np.array([
        np.mean(raw_by_slot[p]) if raw_by_slot[p] else 0.0
        for p in range(1, N_PICKS + 1)
    ])
    counts = np.array([len(share_by_slot[p]) for p in range(1, N_PICKS + 1)])
    return share_mean, share_median, raw_mean, counts


# ---- candidate functional forms (p = pick number, 1..60) ----
def f_exp(p, a, k):
    return a * np.exp(-k * p)

def f_power(p, a, b):
    return a * np.power(p, -b)

def f_hybrid(p, a, b, k):
    # power-law decay with an exponential tail — captures steep top + fat mid
    return a * np.power(p, -b) * np.exp(-k * p)


def fit_form(func, p, y, p0):
    popt, _ = curve_fit(func, p, y, p0=p0, maxfev=200000)
    pred = func(p, *popt)
    ss_res = np.sum((y - pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r2 = 1 - ss_res / ss_tot
    return popt, pred, r2


def normalize(v):
    return v / v[0]


def main():
    data = load_picks()
    years = sorted(data.keys())
    share_mean, share_median, raw_mean, counts = per_slot_curves(data)

    p = np.arange(1, N_PICKS + 1, dtype=float)

    # Fit smooth forms to the maturity-neutral mean-share curve.
    fits = {}
    fits["exponential"] = fit_form(f_exp, p, share_mean, p0=(share_mean[0], 0.05))
    fits["power"] = fit_form(f_power, p, share_mean, p0=(share_mean[0], 1.0))
    fits["hybrid"] = fit_form(f_hybrid, p, share_mean, p0=(share_mean[0], 0.8, 0.01))

    best_name = max(fits, key=lambda k: fits[k][2])
    best_popt, best_pred, best_r2 = fits[best_name]

    # Normalize everything to pick 1 = 1.0 for apples-to-apples vs the app curve.
    model_emp = normalize(share_mean)            # empirical (mean share)
    model_emp_med = normalize(share_median)       # empirical (median share, robust)
    model_fit = normalize(best_pred)              # smooth best-fit
    model_raw = normalize(raw_mean)               # naive (maturity-biased)
    app = np.array([APP_CURVE[i] for i in range(1, N_PICKS + 1)])

    # Compare app curve vs the smooth empirical model.
    diff = app - model_fit
    rmse = math.sqrt(np.mean(diff ** 2))
    mae = np.mean(np.abs(diff))
    corr = np.corrcoef(app, model_fit)[0, 1]

    # ---------------- report ----------------
    print("=" * 78)
    print("DRAFT-VALUE CURVE MODEL  —  fitted to historical Win Shares")
    print("=" * 78)
    print(f"Source classes: {years[0]}-{years[-1]}  ({len(years)} draft years)")
    print(f"Picks per slot: {counts.min()}-{counts.max()} observations\n")

    print("Smooth-fit quality on maturity-neutral mean-share curve:")
    for name, (popt, _, r2) in fits.items():
        print(f"  {name:12s}  R^2={r2:.4f}  params={np.round(popt, 5).tolist()}")
    print(f"  --> best: {best_name} (R^2={best_r2:.4f})\n")

    print("App curve vs smooth empirical model (both normalized pick 1 = 1.0):")
    print(f"  Pearson r = {corr:.4f}")
    print(f"  RMSE      = {rmse:.4f}")
    print(f"  MAE       = {mae:.4f}\n")

    print(f"{'Pk':>3} {'APP':>8} {'MODEL':>8} {'EMP-mean':>9} {'EMP-med':>8} "
          f"{'RAW':>7} {'app/model':>9}")
    for i in range(N_PICKS):
        ratio = app[i] / model_fit[i] if model_fit[i] > 0 else float('nan')
        print(f"{i+1:>3} {app[i]:>8.4f} {model_fit[i]:>8.4f} {model_emp[i]:>9.4f} "
              f"{model_emp_med[i]:>8.4f} {model_raw[i]:>7.4f} {ratio:>9.2f}")

    # Where does the app curve most disagree with history?
    print("\nLargest disagreements (app - model), by |diff|:")
    order = np.argsort(-np.abs(diff))
    for j in order[:8]:
        sign = "app RICHER" if diff[j] > 0 else "app CHEAPER"
        print(f"  Pk {j+1:>2}: app={app[j]:.3f}  model={model_fit[j]:.3f}  "
              f"({sign} by {abs(diff[j]):.3f})")

    # ---------------- robustness: metric x class-maturity grid ----------------
    print("\n" + "-" * 78)
    print("ROBUSTNESS — best-fit curve under different metric / class choices")
    print("(all normalized pick 1 = 1.0; compare tail steepness vs APP)")
    print("-" * 78)
    vorp = load_picks(VORP_COL)
    variant_defs = {
        "WS  all (2003-24)": (data, sorted(data)),
        "WS  mature (-2018)": (data, [y for y in data if y <= 2018]),
        "VORP all": (vorp, sorted(vorp)),
        "VORP mature (-2018)": (vorp, [y for y in vorp if y <= 2018]),
    }
    variants_out = {}
    hdr = f"{'variant':22s} {'R2':>5} " + " ".join(f"pk{p:<2}" for p in (10, 20, 30, 45, 60))
    print(hdr)
    print(f"{'APP CURVE':22s} {'-':>5} " +
          " ".join(f"{app[p-1]:>4.3f}" for p in (10, 20, 30, 45, 60)))
    for vname, (vdata, vyears) in variant_defs.items():
        sb = {p: [] for p in range(1, N_PICKS + 1)}
        for y in vyears:
            tot = sum(vdata[y].values())
            if tot <= 0:
                continue
            for pp, vv in vdata[y].items():
                sb[pp].append(vv / tot)
        vshare = np.array([np.mean(sb[pp]) if sb[pp] else 0.0
                           for pp in range(1, N_PICKS + 1)])
        (_, vpred, vr2) = fit_form(f_hybrid, p, vshare, p0=(vshare[0], 0.8, 0.01))
        vnorm = normalize(vpred)
        variants_out[vname] = {str(i + 1): round(float(vnorm[i]), 5)
                               for i in range(N_PICKS)}
        print(f"{vname:22s} {vr2:>5.3f} " +
              " ".join(f"{vnorm[p-1]:>4.3f}" for p in (10, 20, 30, 45, 60)))
    print("\nReading it: WS gives role players longevity credit -> flatter tail "
          "(your curve looks too steep).\nVORP credits only value-above-replacement "
          "-> steeper, closest to your curve.")

    # ---------------- write JSON for the site / further use ----------------
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "source": "public/Draft Valuation.xlsx",
            "metric": "career Win Shares (negatives floored at 0)",
            "classes": [years[0], years[-1]],
            "n_years": len(years),
            "method": "within-year value share, averaged across classes; "
                      "smooth fit normalized so pick 1 = 1.0",
            "best_fit_form": best_name,
            "best_fit_params": np.round(best_popt, 6).tolist(),
            "best_fit_r2": round(best_r2, 4),
            "vs_app_curve": {"pearson_r": round(corr, 4),
                             "rmse": round(rmse, 4), "mae": round(mae, 4)},
        },
        "curves": {
            "model_fit": {str(i + 1): round(float(model_fit[i]), 5) for i in range(N_PICKS)},
            "empirical_mean": {str(i + 1): round(float(model_emp[i]), 5) for i in range(N_PICKS)},
            "empirical_median": {str(i + 1): round(float(model_emp_med[i]), 5) for i in range(N_PICKS)},
            "raw_mean_biased": {str(i + 1): round(float(model_raw[i]), 5) for i in range(N_PICKS)},
            "app_curve": {str(i + 1): round(float(app[i]), 5) for i in range(N_PICKS)},
        },
        "robustness_variants": variants_out,
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2))
    print(f"\nWrote {OUT_JSON.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
